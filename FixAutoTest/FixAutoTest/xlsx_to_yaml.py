#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Excel → YAML 测试用例批量转换工具
用法: python3 xlsx_to_yaml.py [--dry-run]
"""

import os, sys, yaml

# ---- Tag 号 → 可读名称 ----
TAG_TO_NAME = {
    '1':'Account','6':'AvgPx','8':'BeginString','9':'BodyLength','10':'CheckSum',
    '11':'ClOrdID','14':'CumQty','15':'Currency','17':'ExecID','20':'ExecTransType',
    '21':'HandlInst','31':'LastPx','32':'LastQty','34':'MsgSeqNum','35':'MsgType',
    '37':'OrderID','38':'OrderQty','39':'OrdStatus','40':'OrdType','41':'OrigClOrdID',
    '44':'Price','49':'SenderCompID','50':'SenderSubID','52':'SendingTime','54':'Side',
    '55':'Symbol','56':'TargetCompID','57':'TargetSubID','58':'Text','59':'TimeInForce',
    '60':'TransactTime','77':'OpenClose','98':'EncryptMethod','99':'StopPx',
    '108':'HeartBtInt','110':'MinQty','116':'OnBehalfOfSubID','117':'QuoteID',
    '131':'QuoteReqID','132':'BidPx','133':'OfferPx','134':'BidSize','135':'OfferSize',
    '141':'ResetSeqNumFlag','146':'NoRelatedSym','207':'SecurityExchange',
    '262':'MDReqID','263':'SubscriptionRequestType','264':'MarketDepth',
    '265':'MDUpdateType','295':'NoQuoteEntries','296':'NoQuoteSets','298':'QuoteCancelType',
    '299':'QuoteEntryID','302':'QuoteSetID','304':'TotNoQuoteEntries',
    '321':'SecurityRequestType','432':'ExpireDate','448':'PartyID','452':'PartyRole',
    '453':'NoPartyIDs','702':'NoPositions','703':'PosType','704':'LongQty',
    '709':'PosTransType','710':'PosReqID','712':'PosMaintAction','713':'OrigPosReqRefID',
    '714':'PosMaintRptRefID','715':'ClearingBusinessDate','721':'PosMaintRptID',
    '722':'PosMaintRptID2','1151':'SecurityGroup','1166':'QuoteMsgID',
    '20001':'HedgeFlag','20002':'CombDirection','20003':'CombDirection2',
    '20004':'FrontID','20005':'SessionID','20006':'IsSwapOrder','20009':'OptionSelfHedge',
    '20011':'ExecDeclare',
}

MSG_TYPES = {'0':'Heartbeat','A':'Logon','D':'NewOrderSingle','F':'OrderCancelRequest',
    'G':'OrderCancelReplaceRequest','H':'OrderStatusRequest','V':'MarketDataRequest',
    'i':'MassQuote','Z':'QuoteCancel','c':'SecurityDefinitionRequest',
    'AB':'NewOrderMultileg','AL':'PositionMaintenanceRequest',
    '8':'ExecutionReport','AM':'PositionMaintenanceReport','AI':'QuoteStatusReport'}

SKIP_NAMES = {'BeginString','BodyLength','CheckSum','MsgSeqNum','TargetCompID',
    'SenderCompID','SendingTime','TransactTime','EncryptMethod','HeartBtInt',
    'ResetSeqNumFlag','FrontID','SessionID','TargetSubID'}

SKIP_TAGS = {'8','9','10','34','49','50','52','56','57','60','98','108','141','20004','20005'}

def tag_to_name(tag):
    return TAG_TO_NAME.get(str(tag), f'Tag_{tag}')

def is_skip(tag):
    return str(tag) in SKIP_TAGS

def fix_value(name, val):
    """Keep values as-is for now, let YAML loader resolve them later"""
    if isinstance(val, float) and val == int(val):
        return int(val)
    return val

def convert_sheet(wb_path, out_path):
    import openpyxl
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    sheets = wb.sheetnames
    s1 = wb[sheets[0]]
    s2 = wb[sheets[1]]

    # Determine config file
    fname = os.path.basename(wb_path).replace('.xlsx', '')

    # Parse all request steps
    req_steps = []
    for i in range(2, s1.max_row + 1):
        ref = s1.cell(i, 2).value
        row = int(s1.cell(i, 4).value or 0)
        if ref not in sheets:
            continue
        ds = wb[ref]
        tags = {}
        for c in range(1, ds.max_column + 1):
            tag = ds.cell(3, c).value
            val = ds.cell(row + 3, c).value
            if tag is None or val is None or str(val).strip() == '':
                continue
            tag = str(int(tag)) if isinstance(tag, float) else str(tag)
            tags[tag] = fix_value(tag_to_name(tag), val)
        req_steps.append({'sheet': ref, 'tags': tags})

    # Parse all response steps
    rsp_steps = []
    for i in range(2, s2.max_row + 1):
        ref = s2.cell(i, 2).value
        row = int(s2.cell(i, 4).value or 0)
        if ref not in sheets:
            continue
        ds = wb[ref]
        tags = {}
        for c in range(1, ds.max_column + 1):
            tag = ds.cell(3, c).value
            val = ds.cell(row + 3, c).value
            if tag is None or val is None or str(val).strip() == '':
                continue
            tag = str(int(tag)) if isinstance(tag, float) else str(tag)
            if tag == 'keylist':
                continue
            tags[tag] = fix_value(tag_to_name(tag), val)
        # keylist
        kl_cell = ds.cell(row + 3, ds.max_column).value  # typically last column
        rsp_steps.append({'sheet': ref, 'tags': tags})

    # Generate YAML content
    lines = []
    lines.append(f"# 用例: {fname}")
    lines.append(f"# 自动生成自: {os.path.basename(wb_path)}")
    lines.append("")
    lines.append(f"name: \"{fname}\"")
    lines.append("")
    lines.append("steps:")

    for idx, req in enumerate(req_steps):
        rt = req['tags']
        mt = str(rt.get('35', '?'))
        mn = MSG_TYPES.get(mt, f'MsgType_{mt}')
        sym = str(rt.get('55', ''))
        qty = str(rt.get('38', ''))
        px = str(rt.get('44', ''))
        side = str(rt.get('54', ''))
        closer = str(rt.get('77', ''))
        tif = str(rt.get('59', ''))
        acct = str(rt.get('1', ''))

        lines.append(f"  # ---- Step {idx+1}: {mn} ----")
        lines.append(f"  - request:")

        for tag, val in rt.items():
            if is_skip(tag):
                continue
            name = tag_to_name(tag)
            v = str(val).strip()
            if not v or v == '0' and name == 'AvgPx':
                continue
            lines.append(f"      {name}: {v}")

        # Match keys from response
        if idx < len(rsp_steps):
            rsp = rsp_steps[idx]['tags']
            rmt = str(rsp.get('35', '8'))
            rmn = MSG_TYPES.get(rmt, f'MsgType_{rmt}')
            lines.append(f"    expect:")
            lines.append(f"      MsgType: \"{rmt}\"  # {rmn}")
            match_list = []
            for tag, val in rsp.items():
                if is_skip(tag):
                    continue
                name = tag_to_name(tag)
                v = str(val).strip()
                if not v:
                    continue
                lines.append(f"      {name}: {v}")
                # Add to match list (non-system fields)
                if name not in SKIP_NAMES and name not in ('Text','SenderSubID'):
                    match_list.append(name)
            lines.append(f"      match: [{', '.join(match_list)}]")

        lines.append("")

    yaml_content = '\n'.join(lines)

    # Write YAML file
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(yaml_content)
    print(f"  → {out_path}  ({len(req_steps)} steps)")

def main():
    testcase_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'testcase')
    dry_run = '--dry-run' in sys.argv

    for f in sorted(os.listdir(testcase_dir)):
        if not f.endswith('.xlsx'):
            continue
        xlsx_path = os.path.join(testcase_dir, f)
        yaml_path = os.path.join(testcase_dir, f.replace('.xlsx', '.yaml'))

        if dry_run:
            print(f"Would convert: {f} → {os.path.basename(yaml_path)}")
        else:
            print(f"Converting: {f}")
            try:
                convert_sheet(xlsx_path, yaml_path)
            except Exception as e:
                print(f"  ERROR: {e}")

    print(f"\nDone. {'(dry-run)' if dry_run else ''}")

if __name__ == '__main__':
    main()
